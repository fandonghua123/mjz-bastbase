import openpyxl

#写
# book = openpyxl.Workbook()
# sheet = book.active #默认的sheet
# #sheet2 = book.get_sheet_by_name('sheet1')
# # sheet.append( ['id','username','password','error_count'])
# # sheet.append( [1,'wyj','123456',0])
# # sheet.append( [2,'wyj','123456'])
# sheet['a1'] = 'id' #指定行列
# sheet['b1'] = 'username'#
# sheet.cell(3,1,'1')#指定行和列
# book.save('user.xlsx')

book = openpyxl.load_workbook('user.xlsx')

sheet = book.active
# print(sheet.cell(1,1).value)
# print(sheet['a1'].value)
# print(list(sheet.rows)) #所有行的数据
# l = []
# print(list(sheet.columns)) #所有列

# for row in sheet.rows:
#     t = []
#     for col in row:
#         t.append(col.value)
#     l.append(t)
# print(l)

# print(sheet[1:10])#第几行到第几行

sheet.cell(0,0,'sdfsd')
sheet.delete_cols(1)#删除列
sheet.delete_rows(1)#函数行
book.save('user.xlsx')
